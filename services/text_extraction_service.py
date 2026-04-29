"""Best-effort text extraction for AI and search."""
from pathlib import Path


class TextExtractionService:
    def extract(self, file_path: Path, extension: str, mime_type: str) -> str:
        ext = extension.lower().lstrip(".")
        try:
            if ext == "pdf":
                return self._pdf(file_path)
            if ext == "docx":
                return self._docx(file_path)
            if ext in ("xlsx", "xls"):
                return self._excel(file_path)
            if ext == "csv" or mime_type == "text/csv":
                return file_path.read_text(errors="ignore")[:200_000]
            if ext in ("txt",) or mime_type.startswith("text/"):
                return file_path.read_text(errors="ignore")[:200_000]
        except OSError:
            return ""
        return ""

    def _pdf(self, path: Path) -> str:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        parts: list[str] = []
        for page in reader.pages[:50]:
            t = page.extract_text() or ""
            parts.append(t)
        return "\n".join(parts)[:500_000]

    def _docx(self, path: Path) -> str:
        import docx

        d = docx.Document(str(path))
        return "\n".join(p.text for p in d.paragraphs)[:500_000]

    def _excel(self, path: Path) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets[:10]:
            for row in sheet.iter_rows(max_row=500, values_only=True):
                parts.append(" ".join(str(c) for c in row if c is not None))
        wb.close()
        return "\n".join(parts)[:500_000]
